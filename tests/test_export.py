"""Unit tests for tools/export.py -- Excel + JSON export."""
from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.storage import FMEAStorage
from tools.export import (
    RPZ_COLORS,
    export_json,
    export_excel,
    export_fmea,
    _style_header,
    _create_overview_sheet,
    _create_fmea_sheet,
    _create_causes_sheet,
    _create_measures_sheet,
)
from tools._base import STOP_LABELS, STOP_ORDER

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class _ExportTestCase(unittest.TestCase):
    """Base class: creates a temp DB with one full FMEA chain and cleans up."""

    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = FMEAStorage(self.db_path)
        self._seed_data()
        # temp dir for output files
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        self.db.close()
        for suffix in ("", "-shm", "-wal"):
            path = self.db_path + suffix
            if os.path.exists(path):
                os.remove(path)
        # clean output files
        for f in Path(self.tmp_dir).iterdir():
            f.unlink()
        os.rmdir(self.tmp_dir)

    def _seed_data(self):
        """Create a minimal but complete FMEA dataset."""
        self.project_id = self.db.create_project("Test-Projekt", "Testanlage")
        comp_id = self.db.insert_component(
            project_id=self.project_id,
            komp_id="KOMP-001",
            name="Reaktor",
            typ="Apparat",
            kategorie="Verfahrenstechnik",
            system_name="Reaktorsystem",
        )
        func_id = self.db.insert_function(
            component_id=comp_id,
            funktion_id="F-001",
            typ="Prozess",
            beschreibung="Temperatur halten",
        )
        self.fm_id = self.db.insert_failure_mode(
            function_id=func_id,
            fehler_id="FM-001",
            fehlermodus="Temperatur zu hoch",
            fehlerart="Funktionsverlust",
        )
        self.db.insert_failure_effect(
            failure_mode_id=self.fm_id,
            mensch_stufe="hoch",
            mensch_beschreibung="Verbrennungsgefahr",
            umwelt_stufe="mittel",
            umwelt_beschreibung="Emissionen moeglich",
            anlage_stufe="hoch",
            anlage_beschreibung="Beschaedigung Reaktor",
            kosten_stufe="mittel",
            kosten_beschreibung="Reparaturkosten",
        )
        self.db.insert_failure_cause(
            failure_mode_id=self.fm_id,
            ursache_id="U-001",
            beschreibung="Kuehlung ausgefallen",
            herkunft="Betrieb",
            praeventionsphase="Betrieb",
            praeventionshinweis="Redundante Kuehlung",
        )
        self.db.insert_risk_assessment(
            failure_mode_id=self.fm_id,
            S=8, O=5, D=4,
            begruendung_S="Verbrennungsgefahr",
            begruendung_O="Kuehlung selten defekt",
            begruendung_D="Temperaturueberwachung vorhanden",
        )
        self.db.insert_measure(
            failure_mode_id=self.fm_id,
            name="Redundante Kuehlung",
            abe_kategorie="Vermeidung",
            stop_kategorie="T",
            beschreibung="Zweite Kuehlleitung installieren",
            S_neu=8, O_neu=2, D_neu=3,
            begruendung="Ausfallwahrscheinlichkeit sinkt",
            iteration=1,
        )


class _EmptyProjectTestCase(unittest.TestCase):
    """Base class with an empty project (no components)."""

    def setUp(self):
        fd, self.db_path = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.db = FMEAStorage(self.db_path)
        self.project_id = self.db.create_project("Leer-Projekt", "Leere Anlage")
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        self.db.close()
        for suffix in ("", "-shm", "-wal"):
            path = self.db_path + suffix
            if os.path.exists(path):
                os.remove(path)
        for f in Path(self.tmp_dir).iterdir():
            f.unlink()
        os.rmdir(self.tmp_dir)


# ═══════════════════════════════════════════════════════════════
# RPZ_COLORS
# ═══════════════════════════════════════════════════════════════

class TestRPZColors(unittest.TestCase):

    def test_all_four_levels_present(self):
        expected = {"kritisch", "hoch", "mittel", "niedrig"}
        self.assertEqual(set(RPZ_COLORS.keys()), expected)

    def test_colors_are_six_char_hex(self):
        for level, color in RPZ_COLORS.items():
            self.assertRegex(color, r"^[0-9A-F]{6}$",
                             f"RPZ_COLORS['{level}'] is not valid hex: {color}")


# ═══════════════════════════════════════════════════════════════
# JSON Export
# ═══════════════════════════════════════════════════════════════

class TestExportJSON(_ExportTestCase):

    def test_creates_json_file(self):
        out = os.path.join(self.tmp_dir, "report.json")
        result = export_json(self.project_id, out, db_path=self.db_path)
        self.assertEqual(result, out)
        self.assertTrue(os.path.exists(out))

    def test_json_structure(self):
        out = os.path.join(self.tmp_dir, "report.json")
        export_json(self.project_id, out, db_path=self.db_path)
        with open(out, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.assertIn("metadata", data)
        self.assertIn("statistik", data)
        self.assertIn("fmea_data", data)

    def test_metadata_fields(self):
        out = os.path.join(self.tmp_dir, "report.json")
        export_json(self.project_id, out, db_path=self.db_path)
        with open(out, "r", encoding="utf-8") as f:
            meta = json.load(f)["metadata"]
        self.assertEqual(meta["projekt"], "Test-Projekt")
        self.assertEqual(meta["anlage"], "Testanlage")
        self.assertEqual(meta["standard"], "AIAG-VDA FMEA")
        self.assertIn("export_datum", meta)
        self.assertIn("datum", meta)

    def test_fmea_data_not_empty(self):
        out = os.path.join(self.tmp_dir, "report.json")
        export_json(self.project_id, out, db_path=self.db_path)
        with open(out, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.assertGreater(len(data["fmea_data"]), 0)

    def test_creates_parent_directories(self):
        out = os.path.join(self.tmp_dir, "sub", "deep", "report.json")
        export_json(self.project_id, out, db_path=self.db_path)
        self.assertTrue(os.path.exists(out))
        # cleanup nested dirs
        os.unlink(out)
        os.rmdir(os.path.join(self.tmp_dir, "sub", "deep"))
        os.rmdir(os.path.join(self.tmp_dir, "sub"))


class TestExportJSONEmpty(_EmptyProjectTestCase):

    def test_empty_project_produces_valid_json(self):
        out = os.path.join(self.tmp_dir, "empty.json")
        export_json(self.project_id, out, db_path=self.db_path)
        with open(out, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(data["fmea_data"], [])
        self.assertEqual(data["statistik"]["components"], 0)


# ═══════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestExportExcel(_ExportTestCase):

    def test_creates_xlsx_file(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        result = export_excel(self.project_id, out, db_path=self.db_path)
        self.assertEqual(result, out)
        self.assertTrue(os.path.exists(out))

    def test_sheet_names(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        self.assertEqual(wb.sheetnames, ["Übersicht", "FMEA-Analyse", "Fehlerursachen", "Maßnahmen"])
        wb.close()

    def test_overview_sheet_project_info(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Übersicht"]
        self.assertEqual(ws["A1"].value, "FMEA-Report")
        self.assertEqual(ws["A3"].value, "Projekt:")
        self.assertEqual(ws["B3"].value, "Test-Projekt")
        self.assertEqual(ws["A4"].value, "Anlage:")
        self.assertEqual(ws["B4"].value, "Testanlage")
        wb.close()

    def test_fmea_sheet_headers(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["FMEA-Analyse"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 20)]
        self.assertEqual(headers[0], "KOMP-ID")
        self.assertEqual(headers[6], "Fehler-ID")
        self.assertIn("RPZ", headers)
        self.assertIn("RPZ-Status", headers)
        wb.close()

    def test_fmea_sheet_data_row(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["FMEA-Analyse"]
        # row 2 = first data row
        self.assertEqual(ws.cell(row=2, column=1).value, "KOMP-001")
        self.assertEqual(ws.cell(row=2, column=8).value, "Temperatur zu hoch")
        # S=8, O=5, D=4 -> RPZ=160
        self.assertEqual(ws.cell(row=2, column=14).value, 8)
        self.assertEqual(ws.cell(row=2, column=15).value, 5)
        self.assertEqual(ws.cell(row=2, column=16).value, 4)
        self.assertEqual(ws.cell(row=2, column=17).value, 160)
        wb.close()

    def test_causes_sheet_headers(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Fehlerursachen"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        self.assertEqual(headers[0], "Fehler-ID")
        self.assertEqual(headers[3], "Ursache-ID")
        self.assertEqual(headers[4], "Beschreibung")
        wb.close()

    def test_causes_sheet_data(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Fehlerursachen"]
        self.assertEqual(ws.cell(row=2, column=1).value, "FM-001")
        self.assertEqual(ws.cell(row=2, column=4).value, "U-001")
        self.assertEqual(ws.cell(row=2, column=5).value, "Kuehlung ausgefallen")
        wb.close()

    def test_measures_sheet_headers(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Maßnahmen"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
        self.assertEqual(headers[0], "Fehler-ID")
        self.assertEqual(headers[5], "STOP")
        self.assertEqual(headers[6], "Maßnahme")
        self.assertIn("RPZ neu", headers)
        wb.close()

    def test_measures_sheet_stop_label(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Maßnahmen"]
        # stop_kategorie="T" -> label "Technisch"
        self.assertEqual(ws.cell(row=2, column=6).value, "Technisch")
        wb.close()


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestExportExcelEmpty(_EmptyProjectTestCase):

    def test_empty_project_creates_valid_xlsx(self):
        out = os.path.join(self.tmp_dir, "empty.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        self.assertTrue(os.path.exists(out))
        from openpyxl import load_workbook
        wb = load_workbook(out)
        self.assertEqual(len(wb.sheetnames), 4)
        # FMEA-Analyse should have header row only
        ws = wb["FMEA-Analyse"]
        self.assertIsNotNone(ws.cell(row=1, column=1).value)  # header exists
        self.assertIsNone(ws.cell(row=2, column=1).value)  # no data rows
        wb.close()


# ═══════════════════════════════════════════════════════════════
# RPZ Color Mapping in Excel
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestRPZColorMapping(_ExportTestCase):

    def test_rpz_status_cell_has_fill(self):
        out = os.path.join(self.tmp_dir, "report.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["FMEA-Analyse"]
        # RPZ=160 -> status should be "hoch" or "kritisch"
        status_cell = ws.cell(row=2, column=18)
        rpz_status = status_cell.value
        self.assertIn(rpz_status, RPZ_COLORS)
        # Cell should have the matching fill color
        fill_color = status_cell.fill.start_color.rgb
        # openpyxl prepends "00" for alpha in some cases
        self.assertTrue(
            fill_color.endswith(RPZ_COLORS[rpz_status]),
            f"Expected fill ending with {RPZ_COLORS[rpz_status]}, got {fill_color}",
        )
        wb.close()


# ═══════════════════════════════════════════════════════════════
# _style_header
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestStyleHeader(unittest.TestCase):

    def test_applies_bold_white_font(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Test")
        ws.cell(row=1, column=2, value="Header")
        _style_header(ws, 1, 2)
        for col in (1, 2):
            cell = ws.cell(row=1, column=col)
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.font.color.rgb, "00FFFFFF")
        wb.close()

    def test_applies_fill_to_all_columns(self):
        wb = Workbook()
        ws = wb.active
        for c in range(1, 6):
            ws.cell(row=1, column=c, value=f"H{c}")
        _style_header(ws, 1, 5)
        for c in range(1, 6):
            fill = ws.cell(row=1, column=c).fill
            self.assertEqual(fill.fill_type, "solid")
        wb.close()


# ═══════════════════════════════════════════════════════════════
# export_fmea (combined)
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestExportFMEA(_ExportTestCase):

    def test_format_both_returns_two_paths(self):
        base = os.path.join(self.tmp_dir, "combo")
        result = export_fmea(self.project_id, base, db_path=self.db_path, format="both")
        self.assertIn("json", result)
        self.assertIn("excel", result)
        self.assertTrue(result["json"].endswith(".json"))
        self.assertTrue(result["excel"].endswith(".xlsx"))

    def test_format_json_only(self):
        base = os.path.join(self.tmp_dir, "only")
        result = export_fmea(self.project_id, base, db_path=self.db_path, format="json")
        self.assertIn("json", result)
        self.assertNotIn("excel", result)

    def test_format_excel_only(self):
        base = os.path.join(self.tmp_dir, "only")
        result = export_fmea(self.project_id, base, db_path=self.db_path, format="excel")
        self.assertIn("excel", result)
        self.assertNotIn("json", result)

    def test_default_output_path(self):
        result = export_fmea(self.project_id, db_path=self.db_path, format="json")
        self.assertIn("json", result)
        self.assertTrue(os.path.exists(result["json"]))
        # cleanup
        os.unlink(result["json"])


# ═══════════════════════════════════════════════════════════════
# Edge Cases: Missing Fields
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestMissingFields(_ExportTestCase):

    def _add_fm_without_effects(self):
        """Add a second failure mode with no effects and no measures."""
        comp = self.db.get_components(self.project_id)[0]
        func_id = self.db.insert_function(
            component_id=comp["id"],
            funktion_id="F-002",
            typ="Sicherheit",
            beschreibung="Druck begrenzen",
        )
        fm_id = self.db.insert_failure_mode(
            function_id=func_id,
            fehler_id="FM-002",
            fehlermodus="Druck zu hoch",
            fehlerart="Funktionsverlust",
        )
        # risk but no effects, no causes, no measures
        self.db.insert_risk_assessment(
            failure_mode_id=fm_id, S=6, O=3, D=5,
        )
        return fm_id

    def test_fm_without_effects_exports_cleanly(self):
        self._add_fm_without_effects()
        out = os.path.join(self.tmp_dir, "partial.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["FMEA-Analyse"]
        # Should have 2 data rows (header + 2 FMs)
        values = [ws.cell(row=r, column=7).value for r in range(2, 4)]
        self.assertIn("FM-001", values)
        self.assertIn("FM-002", values)
        wb.close()

    def test_fm_without_causes_no_row_in_causes_sheet(self):
        self._add_fm_without_effects()
        out = os.path.join(self.tmp_dir, "partial.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Fehlerursachen"]
        # Only FM-001 has a cause -> only 1 data row
        self.assertEqual(ws.cell(row=2, column=1).value, "FM-001")
        self.assertIsNone(ws.cell(row=3, column=1).value)
        wb.close()

    def test_fm_without_measures_no_row_in_measures_sheet(self):
        self._add_fm_without_effects()
        out = os.path.join(self.tmp_dir, "partial.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Maßnahmen"]
        # Only FM-001 has a measure -> only 1 data row
        self.assertEqual(ws.cell(row=2, column=1).value, "FM-001")
        self.assertIsNone(ws.cell(row=3, column=1).value)
        wb.close()


# ═══════════════════════════════════════════════════════════════
# STOP sorting in measures sheet
# ═══════════════════════════════════════════════════════════════

@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestMeasuresSortOrder(_ExportTestCase):

    def test_measures_sorted_by_stop(self):
        """Multiple measures on one FM should appear in S-T-O-P order."""
        # Add O and S measures (T already exists from seed)
        self.db.insert_measure(
            failure_mode_id=self.fm_id,
            name="Organisatorische Pruefung",
            abe_kategorie="Entdeckung",
            stop_kategorie="O",
            beschreibung="Regelmaessige Inspektion",
            S_neu=8, O_neu=3, D_neu=2,
            iteration=1,
        )
        self.db.insert_measure(
            failure_mode_id=self.fm_id,
            name="Substitution Loesung",
            abe_kategorie="Vermeidung",
            stop_kategorie="S",
            beschreibung="Alternatives Medium",
            S_neu=4, O_neu=2, D_neu=3,
            iteration=1,
        )
        out = os.path.join(self.tmp_dir, "sorted.xlsx")
        export_excel(self.project_id, out, db_path=self.db_path)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Maßnahmen"]
        stop_values = []
        row = 2
        while ws.cell(row=row, column=6).value is not None:
            stop_values.append(ws.cell(row=row, column=6).value)
            row += 1
        # Expected order: Substitution, Technisch, Organisatorisch
        self.assertEqual(stop_values, ["Substitution", "Technisch", "Organisatorisch"])
        wb.close()


if __name__ == "__main__":
    unittest.main()
