"""
FMEA Export -- Generate Excel and JSON reports from the FMEA database.

Usage:
    python tools/export.py <project_id> [output_path]
"""

import json
import sys
from datetime import datetime
from pathlib import Path

if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent.parent))
from config.fmea_standards import D_SCALE, HEADER_COLOR, O_SCALE, RPZ_LABELS, RPZ_THRESHOLDS, S_SCALE
from config.fmea_standards import RPZ_HEX as RPZ_COLORS
from tools._base import STOP_LABELS, _sort_measures_by_stop, tool_entry
from tools.storage import FMEAStorage

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_json(project_id: int, output_path: str, db_path: str = None) -> str:
    """Export full FMEA data as JSON."""
    with FMEAStorage(db_path) as db:
        project = db.get_project(project_id)
        data = db.get_full_fmea_data(project_id)
        stats = db.get_project_statistics(project_id)

    report = {
        "metadata": {
            "projekt": project["name"],
            "anlage": project["anlage_name"],
            "datum": project["datum"],
            "export_datum": datetime.now().isoformat(),
            "standard": "AIAG-VDA FMEA",
        },
        "statistik": stats,
        "fmea_data": data,
    }

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    return output_path


def export_excel(project_id: int, output_path: str, db_path: str = None) -> str:
    """Export FMEA data as formatted Excel workbook."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    with FMEAStorage(db_path) as db:
        project = db.get_project(project_id)
        components = db.get_components(project_id)
        full_data = db.get_full_fmea_data(project_id)
        stats = db.get_project_statistics(project_id)

    wb = Workbook()

    _create_overview_sheet(wb, project, stats, components)
    _create_fmea_sheet(wb, full_data)
    _create_causes_sheet(wb, full_data)
    _create_measures_sheet(wb, full_data)
    _create_legend_sheet(wb)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


ZEBRA_FILL = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") if HAS_OPENPYXL else None
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
) if HAS_OPENPYXL else None


def _apply_zebra_and_border(ws, start_row: int, end_row: int, max_col: int):
    """Apply alternating row shading and thin borders to data rows."""
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if (r - start_row) % 2 == 1 and not cell.fill.start_color.rgb or cell.fill.start_color.rgb == "00000000":
                cell.fill = ZEBRA_FILL


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 55):
    """Auto-fit column widths based on content, within min/max bounds."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            lengths.append(min(len(val), max_width))
        best = max(lengths) + 2 if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def _style_header(ws, row: int, max_col: int):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def _create_overview_sheet(wb, project, stats, components):
    ws = wb.active
    ws.title = "Übersicht"

    ws["A1"] = "FMEA-Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Projekt:"
    ws["B3"] = project["name"]
    ws["A4"] = "Anlage:"
    ws["B4"] = project["anlage_name"] or "N/A"
    ws["A5"] = "Datum:"
    ws["B5"] = project["datum"]
    ws["A6"] = "Status:"
    ws["B6"] = project["status"]

    ws["A8"] = "Statistik"
    ws["A8"].font = Font(bold=True, size=13)
    ws["A9"] = "Komponenten:"
    ws["B9"] = stats["components"]
    ws["A10"] = "Funktionen:"
    ws["B10"] = stats["functions"]
    ws["A11"] = "Fehlermodi:"
    ws["B11"] = stats["failure_modes"]
    ws["A12"] = "Maßnahmen:"
    ws["B12"] = stats["measures"]

    row = 14
    ws.cell(row=row, column=1, value="RPZ-Verteilung").font = Font(bold=True, size=13)
    row += 1
    for status in ["kritisch", "hoch", "mittel", "niedrig"]:
        count = stats["rpz_distribution"].get(status, 0)
        ws.cell(row=row, column=1, value=status.capitalize())
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color=RPZ_COLORS[status], end_color=RPZ_COLORS[status], fill_type="solid"
        )
        if status in ["kritisch", "hoch"]:
            ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Komponenten-Liste").font = Font(bold=True, size=13)
    row += 1
    headers = ["KOMP-ID", "Name", "Typ", "Kategorie", "System"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))
    row += 1

    for comp in components:
        ws.cell(row=row, column=1, value=comp["komp_id"])
        ws.cell(row=row, column=2, value=comp["name"])
        ws.cell(row=row, column=3, value=comp["typ"])
        ws.cell(row=row, column=4, value=comp["kategorie"])
        ws.cell(row=row, column=5, value=comp["system_name"])
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30


def _create_fmea_sheet(wb, full_data):
    ws = wb.create_sheet("FMEA-Analyse")

    headers = [
        "KOMP-ID", "Komponente", "System", "Typ",
        "Funktion-ID", "Funktion",
        "Fehler-ID", "Fehlermodus", "Fehlerart",
        "Mensch", "Umwelt", "Anlage", "Kosten",
        "S", "O", "D", "RPZ", "RPZ-Status",
        "Override"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for entry in full_data:
        effects = entry.get("effects") or {}
        risk = entry.get("risk") or {}

        ws.cell(row=row, column=1, value=entry.get("komp_id", ""))
        ws.cell(row=row, column=2, value=entry.get("komponente", ""))
        ws.cell(row=row, column=3, value=entry.get("system_name", ""))
        ws.cell(row=row, column=4, value=entry.get("komponenten_typ", ""))
        ws.cell(row=row, column=5, value=entry.get("funktion_id", ""))
        ws.cell(row=row, column=6, value=entry.get("funktion_beschreibung", ""))
        ws.cell(row=row, column=7, value=entry.get("fehler_id", ""))
        ws.cell(row=row, column=8, value=entry.get("fehlermodus", ""))
        ws.cell(row=row, column=9, value=entry.get("fehlerart", ""))

        ws.cell(row=row, column=10, value=effects.get("mensch_beschreibung", ""))
        ws.cell(row=row, column=11, value=effects.get("umwelt_beschreibung", ""))
        ws.cell(row=row, column=12, value=effects.get("anlage_beschreibung", ""))
        ws.cell(row=row, column=13, value=effects.get("kosten_beschreibung", ""))

        S = risk.get("S", entry.get("S", ""))
        O = risk.get("O", entry.get("O", ""))
        D = risk.get("D", entry.get("D", ""))
        rpz = risk.get("rpz", entry.get("rpz", ""))
        rpz_status = risk.get("rpz_status", entry.get("rpz_status", ""))

        ws.cell(row=row, column=14, value=S)
        ws.cell(row=row, column=15, value=O)
        ws.cell(row=row, column=16, value=D)
        ws.cell(row=row, column=17, value=rpz)

        status_cell = ws.cell(row=row, column=18, value=rpz_status)
        if rpz_status in RPZ_COLORS:
            status_cell.fill = PatternFill(
                start_color=RPZ_COLORS[rpz_status],
                end_color=RPZ_COLORS[rpz_status],
                fill_type="solid"
            )
            if rpz_status in ["kritisch", "hoch"]:
                status_cell.font = Font(color="FFFFFF", bold=True)

        ws.cell(row=row, column=19, value=risk.get("override_applied", ""))
        row += 1

    _apply_zebra_and_border(ws, 2, row - 1, len(headers))
    _auto_fit_columns(ws)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["H"].width = 40
    ws.auto_filter.ref = ws.dimensions


def _create_causes_sheet(wb, full_data):
    ws = wb.create_sheet("Fehlerursachen")

    headers = [
        "Fehler-ID", "Fehlermodus", "Komponente",
        "Ursache-ID", "Beschreibung", "Herkunft",
        "Präventionsphase", "Präventionshinweis"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for entry in full_data:
        for cause in entry.get("causes", []):
            ws.cell(row=row, column=1, value=entry.get("fehler_id", ""))
            ws.cell(row=row, column=2, value=entry.get("fehlermodus", ""))
            ws.cell(row=row, column=3, value=entry.get("komponente", ""))
            ws.cell(row=row, column=4, value=cause.get("ursache_id", ""))
            ws.cell(row=row, column=5, value=cause.get("beschreibung", ""))
            ws.cell(row=row, column=6, value=cause.get("herkunft", ""))
            ws.cell(row=row, column=7, value=cause.get("praeventionsphase", ""))
            ws.cell(row=row, column=8, value=cause.get("praeventionshinweis", ""))
            row += 1

    _apply_zebra_and_border(ws, 2, row - 1, len(headers))
    _auto_fit_columns(ws)
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["H"].width = 50
    ws.auto_filter.ref = ws.dimensions


def _create_measures_sheet(wb, full_data):
    ws = wb.create_sheet("Maßnahmen")

    headers = [
        "Fehler-ID", "Fehlermodus", "Komponente",
        "RPZ vorher", "Status vorher",
        "STOP", "Maßnahme", "ABE-Kategorie", "Beschreibung",
        "S neu", "O neu", "D neu", "RPZ neu", "Status neu",
        "Begründung", "Iteration"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    stop_fill = {
        "S": PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),
        "T": PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),
        "O": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "P": PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid"),
    }

    row = 2
    for entry in full_data:
        risk = entry.get("risk") or {}
        measures = _sort_measures_by_stop(entry.get("measures", []))
        for measure in measures:
            ws.cell(row=row, column=1, value=entry.get("fehler_id", ""))
            ws.cell(row=row, column=2, value=entry.get("fehlermodus", ""))
            ws.cell(row=row, column=3, value=entry.get("komponente", ""))
            ws.cell(row=row, column=4, value=risk.get("rpz", entry.get("rpz", "")))
            ws.cell(row=row, column=5, value=risk.get("rpz_status", entry.get("rpz_status", "")))

            stop_kat = measure.get("stop_kategorie", "")
            stop_cell = ws.cell(row=row, column=6,
                                value=STOP_LABELS.get(stop_kat, stop_kat))
            if stop_kat in stop_fill:
                stop_cell.fill = stop_fill[stop_kat]

            ws.cell(row=row, column=7, value=measure.get("name", ""))
            ws.cell(row=row, column=8, value=measure.get("abe_kategorie", ""))
            ws.cell(row=row, column=9, value=measure.get("beschreibung", ""))
            ws.cell(row=row, column=10, value=measure.get("S_neu", ""))
            ws.cell(row=row, column=11, value=measure.get("O_neu", ""))
            ws.cell(row=row, column=12, value=measure.get("D_neu", ""))
            ws.cell(row=row, column=13, value=measure.get("rpz_neu", ""))

            rpz_status_neu = measure.get("rpz_status_neu", "")
            status_cell = ws.cell(row=row, column=14, value=rpz_status_neu)
            if rpz_status_neu in RPZ_COLORS:
                status_cell.fill = PatternFill(
                    start_color=RPZ_COLORS[rpz_status_neu],
                    end_color=RPZ_COLORS[rpz_status_neu],
                    fill_type="solid"
                )

            ws.cell(row=row, column=15, value=measure.get("begruendung", ""))
            ws.cell(row=row, column=16, value=measure.get("iteration", 1))
            row += 1

    _apply_zebra_and_border(ws, 2, row - 1, len(headers))
    _auto_fit_columns(ws)
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["O"].width = 50
    ws.auto_filter.ref = ws.dimensions


def _create_legend_sheet(wb):
    """Create a 'Legende' sheet explaining RPZ colors, scales, and STOP categories."""
    ws = wb.create_sheet("Legende")

    ws["A1"] = "FMEA-Legende"
    ws["A1"].font = Font(bold=True, size=16, color=HEADER_COLOR)

    # RPZ Risk Categories
    row = 3
    ws.cell(row=row, column=1, value="RPZ-Risikokategorien").font = Font(bold=True, size=13)
    row += 1
    headers = ["Kategorie", "RPZ-Schwellwert", "Farbe", "Handlungsempfehlung"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))
    row += 1

    rpz_items = [
        ("Kritisch", f"≥ {RPZ_THRESHOLDS['kritisch']}", "kritisch"),
        ("Hoch", f"≥ {RPZ_THRESHOLDS['hoch']}", "hoch"),
        ("Mittel", f"≥ {RPZ_THRESHOLDS['mittel']}", "mittel"),
        ("Niedrig", f"< {RPZ_THRESHOLDS['mittel']}", "niedrig"),
    ]
    for label, threshold, key in rpz_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=threshold)
        color_cell = ws.cell(row=row, column=3, value="████")
        color_cell.fill = PatternFill(start_color=RPZ_COLORS[key], end_color=RPZ_COLORS[key], fill_type="solid")
        if key in ["kritisch", "hoch"]:
            color_cell.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=4, value=RPZ_LABELS[key])
        row += 1

    # STOP Categories
    row += 1
    ws.cell(row=row, column=1, value="STOP-Kategorien").font = Font(bold=True, size=13)
    row += 1
    stop_items = [
        ("S — Substitution", "D5E8D4", "Gefährdung durch alternative Stoffe/Verfahren eliminieren"),
        ("T — Technisch", "DAE8FC", "Technische Schutzmaßnahme installieren"),
        ("O — Organisatorisch", "FFF2CC", "Organisatorische Maßnahme (Schulung, Verfahren, Prüfplan)"),
        ("P — Persönlich", "F8CECC", "Persönliche Schutzausrüstung (PSA), letzte Verteidigungslinie"),
    ]
    for label, color, desc in stop_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=2, value=desc)
        row += 1

    # ABE Categories
    row += 1
    ws.cell(row=row, column=1, value="ABE-Kategorien").font = Font(bold=True, size=13)
    row += 1
    abe_items = [
        ("A — Vermeidung", "Maßnahme verhindert das Auftreten des Fehlers"),
        ("B — Entdeckung", "Maßnahme verbessert die Entdeckbarkeit des Fehlers"),
        ("E — Abschwächung", "Maßnahme reduziert die Auswirkung des Fehlers"),
    ]
    for label, desc in abe_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)
        row += 1

    # S/O/D Scales
    for scale_name, scale_label, scale_data in [("S", "Bedeutung (Severity)", S_SCALE),
                                                  ("O", "Auftreten (Occurrence)", O_SCALE),
                                                  ("D", "Entdeckung (Detection)", D_SCALE)]:
        row += 1
        ws.cell(row=row, column=1, value=f"{scale_label} — Skala 1-10").font = Font(bold=True, size=13)
        row += 1
        scale_headers = ["Wert", "Einstufung", "Beschreibung"]
        for i, h in enumerate(scale_headers, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, len(scale_headers))
        row += 1
        for val in range(1, 11):
            label, desc = scale_data[val]
            ws.cell(row=row, column=1, value=val).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=desc)
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 45


@tool_entry
def export_fmea(project_id: int, output_path: str = None, db_path: str = None,
                format: str = "both") -> dict:
    """
    Export FMEA data. Returns dict with file paths.
    format: 'excel', 'json', or 'both'
    """
    if output_path is None:
        output_dir = Path(__file__).parent.parent / ".tmp"
        output_dir.mkdir(parents=True, exist_ok=True)
        base = f"fmea_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        output_path = str(output_dir / base)

    results = {}
    base = str(Path(output_path).with_suffix(""))

    if format in ("json", "both"):
        json_path = export_json(project_id, f"{base}.json", db_path)
        results["json"] = json_path

    if format in ("excel", "both"):
        excel_path = export_excel(project_id, f"{base}.xlsx", db_path)
        results["excel"] = excel_path

    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python tools/export.py <project_id> [output_path]")
        print("\nExports FMEA data as Excel (.xlsx) and JSON (.json)")
        sys.exit(1)

    pid = int(sys.argv[1])
    out = sys.argv[2] if len(sys.argv) > 2 else None
    result = export_fmea(pid, out)
    print("Export abgeschlossen:")
    for fmt, path in result.items():
        print(f"  {fmt}: {path}")
